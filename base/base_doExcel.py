# -*-coding:utf-8-*-
import openpyxl
import os
from base.base_log import Log
from pprint import pprint


# excel操作，简单的读写
class DoExcel:
    def __init__(self, file_path, sheet_name):
        # openpyxl只能处理xlsx文件，切记！！！
        self.filepath = os.path.abspath(os.path.dirname(os.path.dirname(__file__))) + r'\testdata\{}'.format(file_path)
        self.sheet_name = sheet_name

    # 读取excel内容
    def read_data(self):
        # 打开excel表格
        try:
            workbook = openpyxl.load_workbook(self.filepath)
            # 获取工作表对象
            sheet = workbook[self.sheet_name]
            test_data = []
            # 逐行读取数据，存入列表
            for i in range(2, sheet.max_row+1):
                test_dict = {}
                test_dict['case_id'] = sheet.cell(i, 1).value
                test_dict['case_name'] = sheet.cell(i, 2).value
                test_dict['method'] = sheet.cell(i, 3).value
                test_dict['url'] = sheet.cell(i, 4).value
                test_dict['headers'] = sheet.cell(i, 5).value
                test_dict['param'] = sheet.cell(i, 6).value
                test_dict['expected_1'] = sheet.cell(i, 7).value
                test_dict['expected_2'] = sheet.cell(i, 8).value
                test_dict['response_body'] = sheet.cell(i, 9).value
                test_dict['result'] = sheet.cell(i, 10).value
                test_dict['rely'] = sheet.cell(i, 11).value
                test_data.append(test_dict)
            return test_data
        except Exception as e:
            Log().error('打开excel报错! {}'.format(e))

    # 写入excel数据,参数为行、列、值
    def write_data(self, row, column, value):
        try:
            workbook = openpyxl.load_workbook(self.filepath)
            sheet = workbook[self.sheet_name]
            sheet.cell(row=row, column=column).value = value
            workbook.save(self.filepath)
        except Exception as e:
            Log().error('打开excel报错! {}'.format(e))


if __name__ == '__main__':
    A = DoExcel('test_post_feed.xlsx', 'test_post_feed')
    pprint(A.read_data())
